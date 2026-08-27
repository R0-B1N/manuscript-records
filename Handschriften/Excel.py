# Author: Leonie Giessler
# Created: 2026-08-26
# Version: 1.0

# benoetigte Bibliotheken importieren
import json
import re
import unicodedata
from datetime import date, datetime

from openpyxl import load_workbook

# Sprachcodes
# Zuordnung nach ISO 639-2/B
SPRACHCODES = {
    # Deutsch
    "deutsch": "ger",
    "german": "ger",
    "ger": "ger",
    "deu": "ger",
    "neuhochdeutsch": "ger",
    "fruhneuhochdeutsch": "ger",

    # Historische deutsche Sprachstufen
    "mittelhochdeutsch": "gmh",
    "althochdeutsch": "goh",
    "mittelniederdeutsch": "gml",
    "niederdeutsch": "nds",

    # Latein
    "latein": "lat",
    "lateinisch": "lat",
    "latin": "lat",
    "lat": "lat",

    # Franzoesisch
    "franzosisch": "fre",
    "french": "fre",
    "fre": "fre",
    "fra": "fre",
    "altfranzosisch": "fro",
    "mittelfranzosisch": "frm",

    # Englisch
    "englisch": "eng",
    "english": "eng",
    "eng": "eng",
    "altenglisch": "ang",
    "mittelenglisch": "enm",

    # Italienisch
    "italienisch": "ita",
    "italian": "ita",
    "ita": "ita",

    # Spanisch
    "spanisch": "spa",
    "spanish": "spa",
    "spa": "spa",

    # Portugiesisch
    "portugiesisch": "por",
    "portuguese": "por",
    "por": "por",

    # Niederlaendisch
    "niederlandisch": "dut",
    "hollandisch": "dut",
    "dutch": "dut",
    "dut": "dut",
    "nld": "dut",

    # Griechisch
    "griechisch": "gre",
    "neugriechisch": "gre",
    "gre": "gre",
    "ell": "gre",
    "altgriechisch": "grc",

    # Hebraeisch
    "hebraisch": "heb",
    "hebrew": "heb",
    "heb": "heb",

    # Jiddisch
    "jiddisch": "yid",
    "yiddisch": "yid",
    "yiddish": "yid",
    "yid": "yid",

    # Arabisch
    "arabisch": "ara",
    "arabic": "ara",
    "ara": "ara",

    # Russisch
    "russisch": "rus",
    "russian": "rus",
    "rus": "rus",

    # Polnisch
    "polnisch": "pol",
    "polish": "pol",
    "pol": "pol",

    # Tschechisch
    "tschechisch": "cze",
    "czech": "cze",
    "cze": "cze",
    "ces": "cze",

    # Slowakisch
    "slowakisch": "slo",
    "slovak": "slo",
    "slo": "slo",
    "slk": "slo",

    # Slowenisch
    "slowenisch": "slv",
    "slovenian": "slv",
    "slv": "slv",

    # Kroatisch
    "kroatisch": "hrv",
    "croatian": "hrv",
    "hrv": "hrv",

    # Serbisch
    "serbisch": "srp",
    "serbian": "srp",
    "srp": "srp",

    # Ungarisch
    "ungarisch": "hun",
    "hungarian": "hun",
    "hun": "hun",

    # Daenisch
    "danisch": "dan",
    "danish": "dan",
    "dan": "dan",

    # Schwedisch
    "schwedisch": "swe",
    "swedish": "swe",
    "swe": "swe",

    # Norwegisch
    "norwegisch": "nor",
    "norwegian": "nor",
    "nor": "nor",

    # Rumaenisch
    "rumanisch": "rum",
    "romanian": "rum",
    "rum": "rum",
    "ron": "rum",

    # Tuerkisch
    "turkisch": "tur",
    "turkish": "tur",
    "tur": "tur",

    # Persisch
    "persisch": "per",
    "persian": "per",
    "per": "per",
    "fas": "per",

    # Chinesisch
    "chinesisch": "chi",
    "chinese": "chi",
    "chi": "chi",
    "zho": "chi",

    # Ukrainisch
    "ukrainisch": "ukr",
    "ukrainian": "ukr",
    "ukr": "ukr",

    # Katalanisch
    "katalanisch": "cat",
    "catalan": "cat",
    "cat": "cat",
}

def _normalisiere_sprache(sprache):
  """
  Normalisiert eine Sprachbezeichnung fuer den Vergleich.

  Beispiele:
      "Französisch" -> "franzosisch"
      "HEBRÄISCH"   -> "hebraisch"
  """
  if sprache is None:
    return ""

  text = str(sprache).strip().lower()
  text = text.replace("ß", "ss")

  # Umlaute und andere Akzente entfernen
  text = unicodedata.normalize("NFKD", text)
  text = "".join(
      zeichen
      for zeichen in text
      if not unicodedata.combining(zeichen)
  )

  text = re.sub(r"\s+", " ", text)

  return text.strip()

def sprachen_zu_codes(sprache):
  """
  Wandelt den Klartext einer Sprache in ISO 639-2/B um.

  Bei unbekannter Sprache wird ein leerer String zurueckgegeben.
  """
  if sprache is None:
    return [], []

  text = str(sprache).strip()

  if not text:
    return [], []

  # Verschiedene moegliche Trenner vereinheitlichen
  teile = re.split(
      r"\s*(?:/|;|,|\+|\bund\b|\bu\.\b)\s*",
      text,
      flags=re.IGNORECASE
  )

  codes = []
  unbekannt = []

  for teil in teile:
    teil = teil.strip()

    if not teil:
      continue

    normalisiert = _normalisiere_sprache(teil)
    code = SPRACHCODES.get(normalisiert)

    if code:
      # Doppelte Sprachcodes vermeiden
      if code not in codes:
        codes.append(code)
    else:
      unbekannt.append(teil)

  return codes, unbekannt

# Hilfsfunktionen
def _zelltext(wert):
  """
  Wandelt einen Excel-Zellwert sicher in Text um.
  """
  if wert is None:
    return ""

  if isinstance(wert, float) and wert.is_integer():
    return str(int(wert))

  return str(wert).strip()


def _verfasser_aufteilen(name):
  """
  Zerlegt einen MODS-Namen wie "Brentano, Clemens (1778-1842)"
  in Nachname, Vorname und Lebensdaten.
  """
  match = re.match(r"^(?P<nachname>[^,]+),\s*(?P<rest>.+)$", name)
  if not match:
    return name.strip(), "", ""

  nachname = match["nachname"].strip()
  rest = match["rest"].strip()

  datum_match = re.search(r"\(([^)]*)\)\s*$", rest)
  if datum_match:
    lebensdaten = datum_match.group(1).strip()
    vorname = rest[:datum_match.start()].strip()
  else:
    lebensdaten = ""
    vorname = rest

  return nachname, vorname, lebensdaten

def _jahr_ermitteln(wert):
  """
  Ermittelt aus dem Inhalt von Spalte F das Jahr.

  Beispiele:
      Excel-Datum 12.05.1845 -> 1845
      "1845"                 -> 1845
      "ca. 1845"             -> 1845
      "1845-1847"            -> 1845

  Entspricht damit dem bisherigen Verhalten, bei dem fuer 011@
  nur das erste relevante Jahr verwendet wird.
  """
  if wert is None:
    return ""

  # Echte Excel-Datumszelle
  if isinstance(wert, (datetime, date)):
    return str(wert.year)

  # Reines Zahlenfeld, z.B. 1845
  if isinstance(wert, (int, float)):
    if 0 < wert < 10000:
      return str(int(wert))

  text = str(wert).strip()

  if not text:
    return ""

  # Erstes vierstelliges Jahr suchen
  match = re.search(r"(?<!\d)(\d{4})(?!\d)", text)

  if match:
    return match.group(1)

  # Falls kein vierstelliges Jahr gefunden wurde
  return text[:4]


def Excel_Einlesen(Dateiname, Tabellenblatt=None):
  """
  Liest die Excel-Datei ein.

  Spalten:
      A = Signatur
      C = Hauptverfasser
      D = Haupttitel
      F = Erscheinungsdatum
      G = Schriftsprache

  Nur Zeilen mit einem Haupttitel in Spalte D werden verarbeitet.
  """
  Arbeitsmappe = load_workbook(Dateiname, data_only=True)

  if Tabellenblatt:
    if Tabellenblatt not in Arbeitsmappe.sheetnames:
      raise ValueError(f"Tabellenblatt '{Tabellenblatt}' existiert nicht.")

    Tabelle = Arbeitsmappe[Tabellenblatt]
  else:
    Tabelle = Arbeitsmappe.active

  Ergebnisse = []

  # Zeile 1 = Kopfzeile, daher Beginn bei Zeile 2
  for zeilennummer in range(2, Tabelle.max_row + 1):

    # D = Haupttitel
    titel = _zelltext(
        Tabelle.cell(
            row=zeilennummer,
            column=4
        ).value
    )

    # Leerzeilen und Zusatzzeilen ohne Haupttitel ignorieren
    if not titel:
      continue

    # A = Signatur
    signatur = _zelltext(
        Tabelle.cell(
            row=zeilennummer,
            column=1
        ).value
    )

    # C = Hauptverfasser
    hauptverfasser = _zelltext(
        Tabelle.cell(
            row=zeilennummer,
            column=3
        ).value
    )

    # F = Erscheinungsdatum
    erscheinungsdatum = Tabelle.cell(
        row=zeilennummer,
        column=6
    ).value

    jahr = _jahr_ermitteln(erscheinungsdatum)

    # G = Schriftsprache
    sprache_text = _zelltext(
        Tabelle.cell(
            row=zeilennummer,
            column=7
        ).value
    )

    sprache_codes, unbekannte_sprachen = sprachen_zu_codes(sprache_text)

    if unbekannte_sprachen:
      print(
          f"Warnung in Excel-Zeile {zeilennummer}: "
          f"Folgende Sprachangaben wurden nicht erkannt: "
          f"{', '.join(unbekannte_sprachen)}"
      )

    data = {
      "Titel": titel,
      "Hauptverfasser_Name": hauptverfasser,
      "Jahr": jahr,
      "Signatur": signatur,
      "Sprache_Codes": sprache_codes,
    }

    Ergebnisse.append(data)

  return Ergebnisse


def zu_Pica_Feldern(data):
    """
    Wandelt einen extrahierten Datensatz in ein PICA+-JSON-Record um.
    Format: Liste von Feldern, jedes Feld = [Tag, Occurrence, Code, Wert, ...]
    (siehe https://format.gbv.de/pica/json)

    002@/002C/002D/002E sind feste Werte fuer nicht-autoptisch erschlossene
    Einzelhandschriften (Retrokatalogisierung), siehe interne Vorgabe.
    003@ (PPN) wird bewusst NICHT belegt - die PPN vergibt das Zielsystem
    beim Import. Die Kalliope-ID wird stattdessen im lokalen Feld 0600
    mitgegeben, damit der Bezug zum Quelldatensatz erhalten bleibt.
    """
    Felder = []

    # 002@ - Satzart/Status (fest: Handschrift, einzelne Einheit, retrokatalogisiert)
    Felder.append(["002@", None, "0", "Har"])

    # 002C - Inhaltstyp (fest: Text)
    Felder.append(["002C", None, "a", "Text", "b", "txt"])

    # 002D - Medientyp (fest: ohne Hilfsmittel zu benutzen)
    Felder.append(["002D", None, "a", "ohne Hilfsmittel zu benutzen", "b", "n"])

    # 002E - Datentraegertyp (fest: Band)
    Felder.append(["002E", None, "a", "Band", "b", "nc"])

    # 011@ - Erscheinungs-/Entstehungsjahr
    if data["Jahr"]:
      Felder.append(["011@", None, "a", data["Jahr"]])
    else:
      Felder.append(["011@", None, "a", ""])

    # 013D - Art des Inhalts (GND-Verknuepfung zu "Handschrift")
    # bewusst leer gelassen, PPN-Verlinkung muss noch manuell ergaenzt werden
    Felder.append(["013D", None, "a", "Handschrift"])

    # 010@ - Sprachcode (ISO 639-2/B, z.B. "ger", "fre")
    if data["Sprache_Codes"]:
      feld = ["010@", None]

      for code in data["Sprache_Codes"]:
        feld += ["a", code]

      Felder.append(feld)

    else:
      feld = ["010@", None, "a", "" ]

      Felder.append(feld)

    # 019@ - Erscheinungsland (codiert)
    # ANNAHME: fest auf "XA-DE" gesetzt, da Abfrage auf Institution DE-39
    # eingeschraenkt ist. Keine echte Ableitung aus dem Entstehungsort!
    Felder.append(["019@", None, "a", "XA-DE"])

    # 028A - Hauptverfasser (nur einer, keine weiteren Personen)
    # in Excel keine GND-ID Verweise, deswegen ohne Subfeld 9
    if data["Hauptverfasser_Name"]:
      nachname, vorname, lebensdaten = _verfasser_aufteilen(
          data["Hauptverfasser_Name"])
      feld = ["028A", None, "a", nachname]
      if vorname:
        feld += ["n", vorname]
      if lebensdaten:
        feld += ["d", lebensdaten]
      #if data["Hauptverfasser_GND_ID"]:
      #  feld += ["9", data["Hauptverfasser_GND_ID"]]
      Felder.append(feld)
    else:
      feld = ["028A", None, "a", ""]
      Felder.append(feld)

    # 022A - hier zweckentfremdet fuer die Signatur (statt Werktitel)
    if data["Signatur"]:
      Felder.append(["022A", None, "a", f"Signatur {data['Signatur']}"])

    # 021A - Haupttitel
    if data["Titel"]:
        Felder.append(["021A", None, "a", data["Titel"]])

    # 017C - lokales Verweis-/Sammelfeld (kein offizieller Pica+-Tag!)
    # enthaelt Kalliope-ID
    # hier nicht, da keine Kalliope Abfrage
    #sonstige = []
    #if data["ID"]:
    #    sonstige += ["3", data["ID"]]
    #if sonstige:
    #  Felder.append(["017C", None] + sonstige)

    return Felder

def pica_json_dump(records):
    def feld_zu_zeile(feld):
        return "[ " + ", ".join(json.dumps(teil, ensure_ascii=False) for teil in feld) + " ]"

    zeilen = ["["]
    for i, record in enumerate(records):
        zeilen.append("  [")
        for j, feld in enumerate(record):
            komma = "," if j < len(record) - 1 else ""
            zeilen.append(f"    {feld_zu_zeile(feld)}{komma}")
        komma = "," if i < len(records) - 1 else ""
        zeilen.append(f"  ]{komma}")
    zeilen.append("]")
    return "\n".join(zeilen)

# main
Excel_Dateiname = input("Excel-Datei (.xlsx): ").strip()
Tabellenblatt = input("Tabellenblatt (Enter = aktives Tabellenblatt): ")
Dateiname = input("Dateiname fuer Export (Enter = excel_pica.json): ").strip() or "excel_pica.json"

try:
    Liste = Excel_Einlesen(Excel_Dateiname, Tabellenblatt or None)
    Pica_Datensaetze = [zu_Pica_Feldern(eintrag) for eintrag in Liste]

    with open(Dateiname, "w", encoding="utf-8") as f:
        f.write(pica_json_dump(Pica_Datensaetze))

    print(f"{len(Pica_Datensaetze)} Datensaetze nach '{Dateiname}' exportiert.")
    print(f"Auftrag abgeschlossen!")

except Exception as e:
    print(f"Fehler: {e}")